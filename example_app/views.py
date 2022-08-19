import openpyxl
from django.http import HttpResponse


def excel_formula(request):
    wb = openpyxl.load_workbook('Book1.xlsx')
    ws = wb['Sheet1']
    print('##########')
    print(ws['A1'].data_type)
    print(ws['A1'].value)
    print('##########')
    print('##########')
    print(ws['G1'].data_type)
    print(ws['G1'].value)
    print('##########')
    print('##########')
    print(ws['F1'].data_type)
    print(ws['F1'].value)
    print('##########')
    return HttpResponse()
